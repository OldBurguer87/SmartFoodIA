from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalog import Category, Product, Store


EXPECTED_HEADERS = {
    "Categoria (Consumer)",
    "Código PDV",
    "Produto",
    "Preço",
    "Descrição",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: Any) -> str:
    return clean_text(value)


@dataclass(frozen=True)
class ParsedCatalogRow:
    source_row: int
    enabled: bool
    category: str
    external_code: str
    name: str
    price: Decimal
    description: str | None


@dataclass(frozen=True)
class ImportIssue:
    source_row: int | None
    external_code: str | None
    issue_type: str
    message: str


@dataclass
class ImportReport:
    file_name: str
    store_id: str
    rows_read: int = 0
    rows_valid: int = 0
    categories_created: int = 0
    products_created: int = 0
    products_updated: int = 0
    products_unchanged: int = 0
    products_deactivated: int = 0
    duplicates_ignored: int = 0
    conflicts_skipped: int = 0
    invalid_rows: int = 0
    issues: list[ImportIssue] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["issues"] = [asdict(issue) for issue in self.issues]
        return data

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(self.to_dict(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


class ConsumerCatalogWorkbookParser:
    def parse(self, file_path: Path) -> tuple[list[ParsedCatalogRow], list[ImportIssue], int]:
        if not file_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        worksheet = workbook.active
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(value) for value in raw_headers]
        header_index = {header: index for index, header in enumerate(headers)}

        missing = EXPECTED_HEADERS.difference(header_index)
        if missing:
            raise ValueError(
                "Planilha Consumer sem as colunas obrigatórias: "
                + ", ".join(sorted(missing))
            )

        rows: list[ParsedCatalogRow] = []
        issues: list[ImportIssue] = []
        total_rows = 0

        for source_row, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            total_rows += 1
            try:
                parsed = self._parse_row(values, source_row, header_index)
            except ValueError as exc:
                issues.append(
                    ImportIssue(
                        source_row=source_row,
                        external_code=clean_text(
                            values[header_index["Código PDV"]]
                            if header_index["Código PDV"] < len(values)
                            else ""
                        )
                        or None,
                        issue_type="invalid_row",
                        message=str(exc),
                    )
                )
                continue
            rows.append(parsed)

        return rows, issues, total_rows

    @staticmethod
    def _parse_row(
        values: tuple[Any, ...],
        source_row: int,
        header_index: dict[str, int],
    ) -> ParsedCatalogRow:
        def value(column: str) -> Any:
            index = header_index[column]
            return values[index] if index < len(values) else None

        external_code = clean_text(value("Código PDV"))
        name = clean_text(value("Produto"))
        category = clean_text(value("Categoria (Consumer)")) or "Sem categoria"
        description = clean_text(value("Descrição")) or None
        enabled_value = values[0] if values else True
        enabled = bool(enabled_value) if enabled_value is not None else True

        if not external_code:
            raise ValueError("Código PDV vazio.")
        if not name:
            raise ValueError("Nome do produto vazio.")

        raw_price = value("Preço")
        try:
            price = Decimal(str(raw_price)).quantize(Decimal("0.01"))
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError(f"Preço inválido: {raw_price!r}.") from None

        if price < 0:
            raise ValueError("Preço negativo.")

        return ParsedCatalogRow(
            source_row=source_row,
            enabled=enabled,
            category=category,
            external_code=external_code,
            name=name,
            price=price,
            description=description,
        )


class ConsumerCatalogImportService:
    def __init__(
        self,
        parser: ConsumerCatalogWorkbookParser | None = None,
    ) -> None:
        self.parser = parser or ConsumerCatalogWorkbookParser()

    def import_workbook(
        self,
        db: Session,
        *,
        store_id: UUID,
        file_path: Path,
        report_path: Path | None = None,
    ) -> ImportReport:
        store = db.get(Store, store_id)
        if store is None:
            raise ValueError(f"Loja não encontrada: {store_id}")

        parsed_rows, parsing_issues, rows_read = self.parser.parse(file_path)
        report = ImportReport(
            file_name=file_path.name,
            store_id=str(store_id),
            rows_read=rows_read,
            invalid_rows=len(parsing_issues),
            issues=list(parsing_issues),
        )

        selected_rows = self._deduplicate(parsed_rows, report)
        report.rows_valid = len(selected_rows)

        categories = {
            category.name: category
            for category in db.scalars(
                select(Category).where(Category.store_id == store_id)
            ).all()
        }
        products = {
            product.external_code: product
            for product in db.scalars(
                select(Product).where(Product.store_id == store_id)
            ).all()
        }

        try:
            for row in selected_rows:
                category = categories.get(row.category)
                if category is None:
                    category = Category(
                        store_id=store_id,
                        name=row.category,
                        display_order=len(categories),
                        active=True,
                    )
                    db.add(category)
                    db.flush()
                    categories[row.category] = category
                    report.categories_created += 1

                product = products.get(row.external_code)
                if product is None:
                    product = Product(
                        store_id=store_id,
                        category_id=category.id,
                        external_code=row.external_code,
                        name=row.name,
                        description=row.description,
                        price=row.price,
                        active=row.enabled,
                        available_for_delivery=row.enabled,
                        available_for_takeout=row.enabled,
                    )
                    db.add(product)
                    products[row.external_code] = product
                    report.products_created += 1
                    continue

                changed = self._apply_updates(product, row, category.id)
                if changed:
                    report.products_updated += 1
                else:
                    report.products_unchanged += 1

            # Produtos que existiam no SmartFoodIA, mas desapareceram
            # da nova exportação do Consumer, são desativados.
            imported_codes = {row.external_code for row in selected_rows}

            # Protege códigos encontrados em linhas inválidas ou conflitos,
            # evitando desativação acidental.
            protected_codes = {
                issue.external_code
                for issue in report.issues
                if issue.external_code
            }

            # Segurança: se nenhuma linha válida for importada,
            # não desativa o catálogo inteiro.
            if imported_codes:
                for external_code, product in products.items():
                    if external_code in imported_codes:
                        continue
                    if external_code in protected_codes:
                        continue

                    was_available = (
                        product.active
                        or product.available_for_delivery
                        or product.available_for_takeout
                    )

                    product.active = False
                    product.available_for_delivery = False
                    product.available_for_takeout = False

                    if was_available:
                        report.products_deactivated += 1

            db.commit()
        except Exception:
            db.rollback()
            raise

        if report_path is not None:
            report.save(report_path)

        return report

    @staticmethod
    def _deduplicate(
        rows: list[ParsedCatalogRow],
        report: ImportReport,
    ) -> list[ParsedCatalogRow]:
        grouped: dict[str, list[ParsedCatalogRow]] = {}
        for row in rows:
            grouped.setdefault(row.external_code, []).append(row)

        selected: list[ParsedCatalogRow] = []
        for external_code, candidates in grouped.items():
            first = candidates[0]
            conflicts = [
                candidate
                for candidate in candidates[1:]
                if (
                    candidate.name.casefold() != first.name.casefold()
                    or candidate.price != first.price
                    or candidate.category.casefold() != first.category.casefold()
                )
            ]

            if conflicts:
                report.conflicts_skipped += 1
                report.issues.append(
                    ImportIssue(
                        source_row=first.source_row,
                        external_code=external_code,
                        issue_type="external_code_conflict",
                        message=(
                            "Código PDV repetido com nome, preço ou categoria diferente. "
                            "Nenhuma linha desse código foi importada."
                        ),
                    )
                )
                continue

            report.duplicates_ignored += len(candidates) - 1
            selected.append(first)

        return selected

    @staticmethod
    def _apply_updates(
        product: Product,
        row: ParsedCatalogRow,
        category_id: UUID,
    ) -> bool:
        changes = {
            "category_id": category_id,
            "name": row.name,
            "description": row.description,
            "price": row.price,
            "active": row.enabled,
            "available_for_delivery": row.enabled,
            "available_for_takeout": row.enabled,
        }
        changed = False
        for attribute, new_value in changes.items():
            if getattr(product, attribute) != new_value:
                setattr(product, attribute, new_value)
                changed = True
        return changed
