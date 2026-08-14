from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalog import Product, ProductFamily, Store
from app.services.catalog.importer import clean_text


@dataclass(frozen=True)
class ConsumerFamilyDefinition:
    external_code: str
    name: str
    description: str | None
    selection_name: str
    selection_required: bool
    child_external_codes: tuple[str, ...]


@dataclass
class ConsumerFamilyImportReport:
    families_found: int = 0
    families_created: int = 0
    families_updated: int = 0
    families_deactivated: int = 0
    product_links: int = 0
    child_products_missing: int = 0


class ConsumerFamilyWorkbookParser:
    def parse(
        self,
        file_path: Path,
    ) -> list[ConsumerFamilyDefinition]:
        workbook = load_workbook(
            file_path,
            data_only=True,
            read_only=True,
        )
        worksheet = workbook.active

        raw_headers = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
        )

        headers = [
            clean_text(value)
            for value in raw_headers
        ]
        header_index = {
            header: index
            for index, header in enumerate(headers)
        }

        required = {
            "Código PDV",
            "Produto",
            "Preço",
        }

        missing = required.difference(header_index)

        if missing:
            raise ValueError(
                "Planilha Consumer sem colunas obrigatórias para "
                "famílias: "
                + ", ".join(sorted(missing))
            )

        rows = list(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            )
        )

        definitions: list[ConsumerFamilyDefinition] = []

        def value(row, column):
            index = header_index.get(column)
            if index is None or index >= len(row):
                return None
            return row[index]

        for position, row in enumerate(rows):
            external_code = clean_text(
                value(row, "Código PDV")
            )
            raw_price = value(row, "Preço")

            # No export do Consumer, os agrupadores aparecem como
            # Pxx e não possuem preço. Eles NÃO são códigos de item
            # que possam ser enviados em um pedido.
            if (
                not external_code.startswith("P")
                or raw_price not in (None, "")
            ):
                continue

            name = clean_text(
                value(row, "Produto")
            )

            if not name:
                continue

            description = (
                clean_text(value(row, "Descrição"))
                or None
            )

            parent_context = clean_text(
                value(row, "Categoria (iFood)")
            )

            children: list[str] = []
            selection_labels: list[str] = []

            for next_row in rows[position + 1:]:
                child_code = clean_text(
                    value(next_row, "Código PDV")
                )
                child_price = value(
                    next_row,
                    "Preço",
                )
                child_context = clean_text(
                    value(next_row, "Categoria (iFood)")
                )

                # Começou a próxima família.
                if (
                    child_code.startswith("P")
                    and child_price in (None, "")
                ):
                    break

                is_variation = (
                    child_price not in (None, "")
                    and parent_context
                    and child_context.startswith(
                        parent_context + " - "
                    )
                )

                if is_variation:
                    if not child_code:
                        raise ValueError(
                            f"Família {external_code} possui "
                            "variação sem Código PDV."
                        )

                    children.append(child_code)

                    selection_labels.append(
                        child_context[
                            len(parent_context) + 3:
                        ].strip()
                    )
                    continue

                # Depois que começaram as variações, qualquer item
                # fora do bloco encerra a família.
                if children:
                    break

                # Produto vendável comum imediatamente depois:
                # não pertence a esta família.
                if child_price not in (None, ""):
                    break

            children = list(
                dict.fromkeys(children)
            )

            if not children:
                raise ValueError(
                    f"Família Consumer {external_code} "
                    f"({name}) não possui variações vendáveis."
                )

            label = next(
                (
                    item
                    for item in selection_labels
                    if item
                ),
                "Opção",
            )

            required_selection = any(
                "obrigat" in item.casefold()
                for item in selection_labels
            )

            selection_name = re.sub(
                r"\bObrigat[oó]rio\b|\bOpcional\b",
                "",
                label,
                flags=re.IGNORECASE,
            ).strip(" -")

            if not selection_name:
                selection_name = "Opção"

            definitions.append(
                ConsumerFamilyDefinition(
                    external_code=external_code,
                    name=name,
                    description=description,
                    selection_name=selection_name,
                    selection_required=required_selection,
                    child_external_codes=tuple(children),
                )
            )

        return definitions


class ConsumerFamilyImportService:
    def __init__(
        self,
        parser: ConsumerFamilyWorkbookParser | None = None,
    ) -> None:
        self.parser = (
            parser
            or ConsumerFamilyWorkbookParser()
        )

    def import_workbook(
        self,
        db: Session,
        *,
        store_id: UUID,
        file_path: Path,
    ) -> ConsumerFamilyImportReport:
        store = db.get(Store, store_id)

        if store is None:
            raise ValueError(
                f"Loja não encontrada: {store_id}"
            )

        definitions = self.parser.parse(file_path)

        report = ConsumerFamilyImportReport(
            families_found=len(definitions)
        )

        products = {
            product.external_code: product
            for product in db.scalars(
                select(Product).where(
                    Product.store_id == store_id
                )
            ).all()
        }

        existing_families = {
            family.external_code: family
            for family in db.scalars(
                select(ProductFamily).where(
                    ProductFamily.store_id == store_id
                )
            ).all()
            if family.external_code
        }

        imported_family_codes = {
            definition.external_code
            for definition in definitions
        }

        # Desativa famílias Consumer antigas que não aparecem
        # mais na exportação atual.
        for code, family in existing_families.items():
            if not code.startswith("P"):
                continue

            if code in imported_family_codes:
                continue

            if family.active:
                family.active = False
                report.families_deactivated += 1

            for product in products.values():
                if product.family_id == family.id:
                    product.family_id = None

        missing_children: list[str] = []

        for order, definition in enumerate(definitions):
            family = existing_families.get(
                definition.external_code
            )

            if family is None:
                family = ProductFamily(
                    store_id=store_id,
                    external_code=definition.external_code,
                    name=definition.name,
                    description=definition.description,
                    selection_name=definition.selection_name,
                    selection_required=(
                        definition.selection_required
                    ),
                    display_order=order,
                    active=True,
                )
                db.add(family)
                db.flush()

                existing_families[
                    definition.external_code
                ] = family

                report.families_created += 1

            else:
                changed = False

                values = {
                    "name": definition.name,
                    "description": definition.description,
                    "selection_name": (
                        definition.selection_name
                    ),
                    "selection_required": (
                        definition.selection_required
                    ),
                    "display_order": order,
                    "active": True,
                }

                for attribute, new_value in values.items():
                    if getattr(family, attribute) != new_value:
                        setattr(
                            family,
                            attribute,
                            new_value,
                        )
                        changed = True

                if changed:
                    report.families_updated += 1

            child_codes = set(
                definition.child_external_codes
            )

            # Remove vínculos antigos que não pertencem mais
            # a esta família.
            for product in products.values():
                if (
                    product.family_id == family.id
                    and product.external_code
                    not in child_codes
                ):
                    product.family_id = None

            for child_code in definition.child_external_codes:
                product = products.get(child_code)

                if product is None:
                    missing_children.append(
                        f"{definition.external_code}->{child_code}"
                    )
                    continue

                product.family_id = family.id
                report.product_links += 1

        report.child_products_missing = len(
            missing_children
        )

        if missing_children:
            raise ValueError(
                "Variações de famílias Consumer não encontradas "
                "no catálogo importado: "
                + ", ".join(missing_children[:20])
            )

        db.flush()

        return report
